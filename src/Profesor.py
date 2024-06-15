import openpyxl
from .materia import Materia


class Profesor:
    def __init__(self, nombre, area):
        self.nombre = nombre
        self.area = area
        self.materias = []  # Inicializamos un array vacío para almacenar las materias

    def agregar_materia(self, nombre,carrera, semestre, dia, hora_inicio, hora_final):
        # Creamos un objeto Materia y lo agregamos al array
        materia = Materia(nombre, carrera, semestre, dia, hora_inicio, hora_final)
        self.materias.append(materia)

    def mostrar_informacion(self):
        print('Nombre: {}, Área: {}'.format(self.nombre,self.area))
        print('Materias:')
        for materia in self.materias:
            materia.imprimir()

    def guardar_en_excel(self, archivo):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Horario"

        # Definir la estructura del horario
        horas = [
            "7:50 A 8:40",
            "8:45 A 9:35",
            "9:35 A 10:25",
            "10:30 A 11:20",
            "11:20 A 12:10",
            "12:15 A 1:10",
            "1:10 A 2:00",
            "2:00 A 2:50",
            "2:55 A 3:45"
        ]
        dias = ["HORA", "LUNES", "MARTES", "MIERCOLES", "JUEVES", "VIERNES", "SABADO"]

        # Escribir los encabezados de las columnas (días de la semana)
        ws.append(dias)

        # Escribir las horas en la primera columna
        for hora in horas:
            ws.append([hora])

        # Rellenar el horario con las materias
        for materia in self.materias:
            # Encuentra la fila correspondiente a la hora de inicio
            for row in ws.iter_rows(min_row=2, max_row=len(horas) + 1, min_col=1, max_col=1):
                if materia.hora_inicio in row[0].value:
                    # Encuentra la columna correspondiente al día
                    col_idx = dias.index(materia.dia) + 1
                    # Concatenar nombre de la materia y carrera
                    valor_celda = "{} - {}".format(materia.nombre, materia.carrera)
                    # Escribe el nombre de la materia en la celda correspondiente
                    ws.cell(row=row[0].row, column=col_idx, value=valor_celda)
                    break

        # Guardar el archivo
        wb.save(archivo)
        print('Información guardada en {}'.format(archivo))
