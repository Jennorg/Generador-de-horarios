# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import Alignment
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import coordinate_to_tuple
from openpyxl.worksheet.cell_range import CellRange
import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, PageBreak, Paragraph, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.colors import yellow, blue
import pdfkit

import os  # Agrega esta línea al inicio de tu archivo

class Horario:
    def __init__(self):
        self.horario = None
    
    def definir_horario(self, horario_dict):
        self.horario = horario_dict

    def obtener_nombre_carrera(self, codigo_carrera):
        archivo_datos_horarios = 'DATOSHORARIOS.xlsx'
        wb = openpyxl.load_workbook(archivo_datos_horarios)
        if 'Carreras' not in wb.sheetnames:
            raise KeyError(f"La hoja 'Carreras' no existe en el archivo {archivo_datos_horarios}.")
        
        hoja_carreras = wb['Carreras']
        for fila in hoja_carreras.iter_rows(min_row=2, values_only=True):
            vacio, codigo, nombre = fila[:3]  # Solo tomar las primeras tres columnas
            if codigo == codigo_carrera:
                return nombre
        return None

    def guardar_en_excel(self, datos):
        archivos_generados = []
        for carrera, semestres in datos.items():
            nombre_carrera = self.obtener_nombre_carrera(carrera)
            if nombre_carrera:
                nombre_archivo = f"{nombre_carrera}.xlsx"
            else:
                nombre_archivo = f"{carrera}.xlsx"
            archivos_generados.append(nombre_archivo)
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            for semestre, secciones in semestres.items():
                if isinstance(secciones, list):
                    for idx, horarios in enumerate(secciones):
                        seccion = f"seccion {idx + 1}"
                        self._crear_hoja(wb, semestre, seccion, horarios)
            wb.save(nombre_archivo)
        for archivo in archivos_generados:
            self.convertir_a_pdf(archivo)

    def _crear_hoja(self, wb, semestre, seccion, horarios):
        sheet = wb.create_sheet(title=f"{semestre} {seccion}")
        horas = [
            "7:50 A 8:40", "8:45 A 9:35", "9:35 A 10:25",
            "10:30 A 11:20", "11:20 A 12:10", "12:15 A 1:10",
            "1:10 A 2:00", "2:00 A 2:50", "2:55 A 3:45"
        ]
        sheet.cell(row=1, column=1, value="Hora")
        for i, hora in enumerate(horas):
            sheet.cell(row=i+2, column=1, value=hora)
            sheet.cell(row=i+2, column=1).alignment = Alignment(horizontal='center', vertical='center')
        dias = ["lunes", "martes", "miercoles", "jueves", "viernes"]
        for i, dia in enumerate(dias):
            sheet.cell(row=1, column=i+2, value=dia.capitalize())
            sheet.cell(row=1, column=i+2).alignment = Alignment(horizontal='center', vertical='center')
        for dia, bloques in horarios.items():
            for bloque, materia in enumerate(bloques):
                if materia is not None:
                    # Definir colores de fondo
                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    blue_fill = PatternFill(start_color="00A8FF", end_color="00A8FF", fill_type="solid")
                    valor_celda = f"{materia['materia']} - {materia['profesor']} - Aula: {materia['aula']} - {materia['modalidad']}"
                    cell = sheet.cell(row=bloque + 2, column=dias.index(dia.lower()) + 2, value=valor_celda)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    if 'Presencial' in valor_celda:
                        cell.fill = yellow_fill
                    elif 'Virtual' in valor_celda:
                        cell.fill = blue_fill               
                else:
                    cell = sheet.cell(row=bloque + 2, column=dias.index(dia.lower()) + 2, value="")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = 25

    def convertir_a_pdf(self, archivo_excel):
        df = pd.read_excel(archivo_excel, sheet_name=None)
        nombre_pdf = archivo_excel.replace('.xlsx', '.pdf')
        doc = SimpleDocTemplate(nombre_pdf, pagesize=A4)

        elements = []

        # Obtener el nombre del archivo sin extensión para determinar el encabezado
        nombre_archivo = os.path.basename(archivo_excel).split('.')[0]

        encabezado = [
            "UNIVERSIDAD NACIONAL EXPERIMENTAL DE GUAYANA",
            "VICERRECTORADO ACADÉMICO",
            "COORDINACIÓN GENERAL DE PREGRADO",
            f"COORDINACIÓN {nombre_archivo}"
        ]

        # Construir la ruta relativa al directorio del script
        current_dir = os.path.dirname(__file__)
        logo_path = os.path.join(current_dir, 'images', 'logo.png')
        
        elements.append(Spacer(1, 0.2 * inch))

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Center', alignment=1))  # Definir estilo 'Center' aquí

        styles['Heading1'].fontSize = 12

        for semestre, df_sheet in df.items():
            if logo_path:
                try:
                    elements.append(Image(logo_path, 1 * inch, 1 * inch))
                except OSError:
                    print(f"No se pudo abrir el logo en la ruta: {logo_path}")

            for line in encabezado:
                elements.append(Paragraph(line, styles['Center']))  # Usar el estilo 'Center' para centrar

            elements.append(Paragraph(semestre, styles['Heading1']))
            data = [df_sheet.columns.values.tolist()] + df_sheet.values.tolist()
            col_widths = [1 * inch] * len(df_sheet.columns)

            adjusted_data = []
            for row in data:
                adjusted_row = []
                for cell in row:
                    if pd.isna(cell):
                        adjusted_row.append("")
                    else:
                        adjusted_row.append(cell)
                adjusted_data.append(adjusted_row)

            table = Table(adjusted_data, colWidths=col_widths)

            style = TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
            ])
            adjusted_table = []
            for row in adjusted_data:
                adjusted_row = []
                for cell in row:
                    if isinstance(cell, str):
                        adjusted_cell = self.wrap_text(cell, 1 * inch, 8)
                    else:
                        adjusted_cell = cell
                    adjusted_row.append(adjusted_cell)
                adjusted_table.append(adjusted_row)

  
            # Aplicar color a las celdas no vacías con condición
            for row_idx, row in enumerate(adjusted_table):
                for col_idx, cell in enumerate(row):
                    if cell:
                        if '' in cell:
                            style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.white)
                        if 'Presencial' in cell:
                            style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.yellow)
                        if ' Virtual ' in cell:
                            style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.lightblue)
                        if row_idx > 0 and adjusted_table[row_idx - 1][col_idx] == cell or adjusted_table[row_idx - 2][col_idx] == cell :
                            style.add('LINEBELOW', (col_idx, row_idx - 1), (col_idx, row_idx - 1), 0, colors.yellow)
                            style.add('LINEABOVE', (col_idx, row_idx), (col_idx, row_idx), 0, colors.yellow)
                            adjusted_table[row_idx][col_idx] = ""

            table = Table(adjusted_table, colWidths=col_widths)
            table.setStyle(style)

            elements.append(table)
            elements.append(PageBreak())

        doc.build(elements)

    def wrap_text(self, text, width, font_size):
        max_chars = int(width / (font_size * 0.5))

        words = text.split()
        lines = []
        current_line = ""

        for word in words:
            if len(current_line + " " + word) <= max_chars:
                current_line += " " + word
            else:
                lines.append(current_line.strip())
                current_line = word
        
        lines.append(current_line.strip())
        
        return "\n".join(lines)

 