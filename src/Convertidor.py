import openpyxl
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
import io
from .profesor import Profesor
from .leer_datos import leer_datos

def preparar_datos():
    dias_semana = ["LUNES", "MARTES", "MIERCOLES", "JUEVES", "VIERNES", "SABADO"]
    datos = leer_datos()
    
    if datos is None:
        raise Exception("Error al leer los datos del archivo Excel.")
    
    materias_data = datos["materias"].to_dict('records')
    profesor1 = Profesor("Juan", "Biología")

    for dia in dias_semana:
        for materia_data in materias_data:
            profesor1.agregar_materia(materia_data["Nombre"], "Ing", 1, dia, "2:00", "2:50")

    return profesor1

def convertir_excel_a_pdf(archivo_excel, archivo_pdf):
    # Cargar el libro de trabajo de Excel
    wb = load_workbook(filename=archivo_excel, read_only=True)

    # Obtener la hoja activa
    ws = wb.active

    # Obtener dimensiones del archivo
    rows = ws.max_row
    cols = ws.max_column

    # Crear un lienzo PDF
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Definir el número de filas y columnas de la tabla
    table_rows = []
    for row in ws.iter_rows(min_row=1, max_row=rows, max_col=cols, values_only=True):
        row_data = []
        for cell in row:
            if isinstance(cell, str):
                row_data.append(cell)  # Si ya es una cadena, mantenerla tal cual
            elif cell is None:
                row_data.append('')  # Si es None, agregar cadena vacía
            else:
                # Convertir a cadena unicode y manejar caracteres especiales
                if isinstance(cell, unicode):
                    row_data.append(cell.encode('utf-8', 'replace').decode('utf-8'))
                else:
                    row_data.append(str(cell))  # Convertir otros tipos a cadena

        table_rows.append(row_data)

    # Crear la tabla
    table = Table(table_rows)

    # Estilo de la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),  # Encabezados de fila
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # Color de texto de encabezados de fila
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Alineación centrada para todo
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Alineación vertical centrada para todo
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),  # Rejilla interna con borde negro
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black)  # Borde de la caja con borde negro
    ])

    table.setStyle(style)

    # Tamaño de la tabla y posición
    table.wrapOn(c, width -80, height -200)
    table.drawOn(c, 30, height - 200)

    # Guardar el PDF generado en el archivo
    c.save()
    buffer.seek(0)
    with open(archivo_pdf, 'wb') as f:
        f.write(buffer.read())

    print('Archivo PDF guardado en:')

def ejecutar_proceso():
    profesor1 = preparar_datos()
    profesor1.mostrar_informacion()
    archivo_excel = 'horario_profesor.xlsx'
    profesor1.guardar_en_excel(archivo_excel)
    archivo_pdf = 'horario_profesor.pdf'
    convertir_excel_a_pdf(archivo_excel, archivo_pdf)




